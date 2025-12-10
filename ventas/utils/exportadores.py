# ================================================================
# =                                                              =
# =              UTILIDADES PARA EXPORTACIÓN                    =
# =                                                              =
# ================================================================
#
# Este archivo contiene funciones utilitarias para exportar
# reportes a diferentes formatos (Excel, PDF, CSV).
#
# FUNCIONALIDADES:
# - Exportación a Excel (XLSX) usando openpyxl
# - Exportación a PDF usando ReportLab
# - Funciones reutilizables para todos los reportes

from django.http import HttpResponse
from decimal import Decimal
from datetime import datetime
import csv

# Intentar importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Intentar importar ReportLab para PDF
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ================================================================
# =              EXPORTACIÓN A EXCEL (XLSX)                     =
# ================================================================

def exportar_a_excel(datos, nombre_archivo, titulo="Reporte"):
    """
    Exporta datos a formato Excel (XLSX).
    
    Args:
        datos: Lista de diccionarios con los datos a exportar
        nombre_archivo: Nombre del archivo sin extensión
        titulo: Título del reporte
        
    Returns:
        HttpResponse: Archivo Excel descargable
    """
    
    if not OPENPYXL_AVAILABLE:
        # Fallback a CSV si openpyxl no está disponible
        return exportar_a_csv(datos, nombre_archivo)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=12)
    title_font = Font(bold=True, size=14, color="000000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Fecha de generación
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].alignment = center_alignment
    ws['A2'].font = Font(size=10, italic=True)
    
    # Encabezados (si hay datos)
    if datos:
        # Obtener claves del primer diccionario como encabezados
        encabezados = list(datos[0].keys())
        
        # Escribir encabezados
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=4, column=col_idx, value=str(encabezado).replace('_', ' ').title())
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        
        # Escribir datos
        for row_idx, fila in enumerate(datos, start=5):
            for col_idx, encabezado in enumerate(encabezados, start=1):
                valor = fila.get(encabezado, '')
                # Convertir Decimal a float para Excel
                if isinstance(valor, Decimal):
                    valor = float(valor)
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = border
                if isinstance(valor, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(encabezados) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    
    # Guardar workbook en respuesta
    wb.save(response)
    
    return response


def exportar_a_csv(datos, nombre_archivo):
    """
    Exporta datos a formato CSV (fallback si Excel no está disponible).
    
    Args:
        datos: Lista de diccionarios con los datos a exportar
        nombre_archivo: Nombre del archivo sin extensión
        
    Returns:
        HttpResponse: Archivo CSV descargable
    """
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.csv"'
    
    writer = csv.writer(response)
    
    if datos:
        # Encabezados
        encabezados = list(datos[0].keys())
        writer.writerow([str(h).replace('_', ' ').title() for h in encabezados])
        
        # Datos
        for fila in datos:
            valores = []
            for encabezado in encabezados:
                valor = fila.get(encabezado, '')
                if isinstance(valor, Decimal):
                    valor = float(valor)
                valores.append(valor)
            writer.writerow(valores)
    
    return response


# ================================================================
# =              EXPORTACIÓN A PDF                              =
# ================================================================

def exportar_a_pdf(datos, nombre_archivo, titulo="Reporte", encabezados=None):
    """
    Exporta datos a formato PDF usando ReportLab.
    
    Args:
        datos: Lista de diccionarios con los datos a exportar
        nombre_archivo: Nombre del archivo sin extensión
        titulo: Título del reporte
        encabezados: Lista de nombres de columnas (opcional, se infiere de datos si no se proporciona)
        
    Returns:
        HttpResponse: Archivo PDF descargable
    """
    
    if not REPORTLAB_AVAILABLE:
        # Fallback a CSV si ReportLab no está disponible
        return exportar_a_csv(datos, nombre_archivo)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    
    # Crear documento PDF
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Contenedor para elementos
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(
        f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabla de datos
    if datos:
        # Obtener encabezados
        if not encabezados:
            encabezados = list(datos[0].keys())
        
        # Preparar datos para la tabla
        tabla_data = []
        
        # Encabezados
        tabla_data.append([
            Paragraph(str(h).replace('_', ' ').title(), styles['Normal'])
            for h in encabezados
        ])
        
        # Datos
        for fila in datos:
            valores = []
            for encabezado in encabezados:
                valor = fila.get(encabezado, '')
                if isinstance(valor, Decimal):
                    valor = f"${valor:,.0f}"
                elif isinstance(valor, (int, float)):
                    valor = str(valor)
                valores.append(Paragraph(str(valor), styles['Normal']))
            tabla_data.append(valores)
        
        # Crear tabla
        tabla = Table(tabla_data, colWidths=[None] * len(encabezados))
        
        # Estilo de la tabla
        tabla.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFD700')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#000000')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(tabla)
    
    # Construir PDF
    doc.build(elements)
    
    return response

